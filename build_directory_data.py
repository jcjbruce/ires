from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import load_workbook

WORKBOOK_PATH = Path('/home/ubuntu/indigenous_research_institutions_global_survey_revised.xlsx')
PROJECT_PATH = Path('/home/ubuntu/indigenous-research-directory')
DATA_DIR = PROJECT_PATH / 'client' / 'src' / 'data'
DATA_DIR.mkdir(parents=True, exist_ok=True)

HEADER_MAP = {
    'Continent': 'continent',
    'Country': 'country',
    'Institution name': 'institutionName',
    'Parent institution': 'parentInstitution',
    'Name language group': 'nameLanguageGroup',
    'Primary theme': 'primaryTheme',
    'Secondary themes': 'secondaryThemes',
    'Website': 'website',
    'Brief description': 'description',
}

CONTINENT_ORDER = [
    'Africa',
    'Asia',
    'Europe',
    'North America',
    'Oceania',
    'South America',
    'Regional',
]

CONTINENT_META = {
    'Africa': {'label': 'Africa', 'color': '#9C6644', 'accent': '#E6CCB2', 'description': 'Institutions rooted in African Indigenous scholarship, land, culture, and governance.'},
    'Asia': {'label': 'Asia', 'color': '#355070', 'accent': '#BBD0FF', 'description': 'Research centers and programs across Asia engaging Indigenous peoples, territories, and knowledge systems.'},
    'Europe': {'label': 'Europe', 'color': '#588157', 'accent': '#D8F3DC', 'description': 'European institutions, including Arctic and Sámi-focused research, with strong policy and cultural research links.'},
    'North America': {'label': 'North America', 'color': '#7F5539', 'accent': '#EDE0D4', 'description': 'A large concentration of Indigenous research institutes, centers, hubs, and programs across Canada and the United States.'},
    'Oceania': {'label': 'Oceania', 'color': '#386641', 'accent': '#DDE5B6', 'description': 'Aboriginal, Torres Strait Islander, Māori, and Pacific-centered institutions with deep community and environmental orientations.'},
    'South America': {'label': 'South America', 'color': '#6B705C', 'accent': '#FFE8D6', 'description': 'Institutions across Latin America and the Caribbean focused on Indigenous territories, languages, rights, and environmental stewardship.'},
    'Regional': {'label': 'Regional', 'color': '#6D597A', 'accent': '#EADFF0', 'description': 'Cross-regional institutions whose work spans multiple continents or transnational Indigenous networks.'},
}

wb = load_workbook(WORKBOOK_PATH, data_only=True)
ws = wb['All institutions']
headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
keys = [HEADER_MAP[h] for h in headers]

rows = []
for row_idx in range(2, ws.max_row + 1):
    values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
    if not any(values):
        continue
    record = {keys[i]: (values[i] or '').strip() if isinstance(values[i], str) else (values[i] if values[i] is not None else '') for i in range(len(keys))}
    record['id'] = f"inst-{row_idx - 1:03d}"
    secondary = record['secondaryThemes']
    if isinstance(secondary, str):
        record['secondaryThemesList'] = [item.strip() for item in secondary.replace(';', ',').split(',') if item.strip()]
    else:
        record['secondaryThemesList'] = []
    rows.append(record)

rows.sort(key=lambda item: (item['institutionName'].lower(), item['country'].lower()))

continent_counts = Counter(item['continent'] for item in rows)
country_counts = Counter(item['country'] for item in rows if item['country'])
primary_theme_counts = Counter(item['primaryTheme'] for item in rows if item['primaryTheme'])
language_counts = Counter(item['nameLanguageGroup'] for item in rows if item['nameLanguageGroup'])

by_continent = defaultdict(list)
for item in rows:
    by_continent[item['continent']].append(item)

payload = {
    'meta': {
        'title': 'Indigenous Research Institutions Directory',
        'subtitle': 'Global survey of university, government, NGO, and Indigenous-led research institutions',
        'totalInstitutions': len(rows),
        'continents': CONTINENT_ORDER,
    },
    'continentMeta': CONTINENT_META,
    'institutions': rows,
    'summary': {
        'continentCounts': dict(continent_counts),
        'topCountries': country_counts.most_common(20),
        'topPrimaryThemes': primary_theme_counts.most_common(20),
        'languageGroups': language_counts.most_common(20),
    },
    'byContinent': {continent: by_continent.get(continent, []) for continent in CONTINENT_ORDER},
}

(PROJECT_PATH / 'data_summary.json').write_text(json.dumps(payload['summary'], indent=2, ensure_ascii=False) + '\n', encoding='utf-8')
(DATA_DIR / 'institutions.json').write_text(json.dumps(payload, indent=2, ensure_ascii=False) + '\n', encoding='utf-8')
(DATA_DIR / 'institutions.ts').write_text(
    '/* Design philosophy: editorial cartography with atlas-style browsing, precise filtering, and calm high-density scholarship. */\n'
    'export const directoryData = ' + json.dumps(payload, indent=2, ensure_ascii=False) + ' as const;\n',
    encoding='utf-8'
)

print(f"Exported {len(rows)} institutions to {DATA_DIR / 'institutions.json'} and {DATA_DIR / 'institutions.ts'}")
print('Continents:', ', '.join(f"{name}={continent_counts.get(name, 0)}" for name in CONTINENT_ORDER))
